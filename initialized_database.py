from openpyxl import Workbook
from datetime import datetime

def initialize_database():

    # 1. 生成文件
    workbook = Workbook()
    workbook_name = str(datetime.now())+".xlsx"

    # 2. 在 Python 文件中生成表头数据

    USER_INFO = ["ID","user_id","user_name","create_time","total_matches","total_games","total_win_games", "total_income"]
    MATCH_INFO = ["ID",	"match_id",	"create_time","small_blind","big_blind","total_games"]
    USER_MATCH_RELATION = ["ID","match_id",	"user_id",	"total_games","total_win_games","total_income","total_chip"]
    GAME_INFO = ["ID","game_id","total_table_chip","table_cards","is_flip","is_turn","is_river","create_time"]
    USER_GAME_RELATION = ["ID","game_id","user_id","user_cards","bet_chip","is_win","total_income","is_allin","is_fold", "position"]

    database_titles = [USER_INFO, MATCH_INFO, USER_MATCH_RELATION, GAME_INFO, USER_GAME_RELATION]

    # 3. 生成多个sheet 并为每个 sheet 填充好表头数据
    for i in range(5):
        worksheet = workbook.create_sheet()
        # print("现在正在处理第{}个表格".format(i+1))
        title_count = len(database_titles[i])
        # print("第{}个表格的字段共有{}个".format((i + 1),title_count))
        for title in range(title_count):
            # print("第{}个表格的字段是{}".format(title+1, database_titles[i][title]))
            worksheet.cell(1,title+1).value = database_titles[i][title]

    workbook.save(workbook_name)
    return workbook_name

