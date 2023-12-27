from openpyxl import load_workbook
from datetime import datetime

# 1. 创建一场比赛
def create_match(database, small_blind=10, big_blind=20):
    workbook = load_workbook(database)
    worksheet = workbook.worksheets[2]
    max_row = worksheet.max_row
    worksheet.cell(max_row + 1, 1).value = max_row
    worksheet.cell(max_row + 1, 2).value = max_row
    worksheet.cell(max_row + 1, 3).value = datetime.now()
    worksheet.cell(max_row + 1, 4).value = small_blind
    worksheet.cell(max_row + 1, 5).value = big_blind
    workbook.save(database)
    workbook.close()
    message = "Match created success"
    return message

def join_match(database):
    user_id = input("请输入用户ID： ")
    match_id = input("请输入matchID： ")
    workbook = load_workbook(database)
    worksheet = workbook.worksheets[3]
    max_row = worksheet.max_row
    for row in worksheet.iter_rows():
        if row[1].value == match_id and row[2].value == user_id:
            return "请勿重复加入"
        else:
            worksheet.cell(max_row + 1,1).value = max_row
            worksheet.cell(max_row + 1, 2).value = match_id
            worksheet.cell(max_row + 1, 3).value = user_id
            worksheet.cell(max_row + 1, 4).value = "True"

    workbook.save(database)
    workbook.close()
    message = "加入比赛成功"
    return message

def left_match(database):
    user_id = input("请输入用户ID： ")
    match_id = input("请输入matchID： ")
    workbook = load_workbook(database)
    worksheet = workbook.worksheets[3]
    for row in worksheet.iter_rows():
        if row[1].value == match_id and row[2].value == user_id:
            row[3].value = "False"
            workbook.save(database)
            workbook.close()
            message = "退出比赛成功"
            return message
        else:
            pass
    return "用户未加入该场比赛"







