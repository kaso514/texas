from datetime import datetime
from openpyxl import load_workbook
from initialized_database import initialize_database
# 1. 用户注册

def register(database):
    user_name = input("请输入用户名称：")
    workbook = load_workbook(database)
    worksheet = workbook.worksheets[1]
    max_row = worksheet.max_row
    worksheet.cell(max_row + 1, 1).value = max_row
    worksheet.cell(max_row + 1, 2).value = max_row
    worksheet.cell(max_row + 1, 3).value = user_name
    worksheet.cell(max_row + 1, 4).value = datetime.now()
    worksheet.cell(max_row + 1, 5).value = 0
    worksheet.cell(max_row + 1, 6).value = 0
    worksheet.cell(max_row + 1, 7).value = 0
    worksheet.cell(max_row + 1, 8).value = 0
    workbook.save(database)
    workbook.close()
    message = "{}已完成注册".format(user_name)
    return message

# history happened
# 2. 查询用户数据 / 获取用户
def get_user_info(user_id,database):
    workbook = load_workbook(database)
    worksheet = workbook.worksheets[1]
    for cell in worksheet["B"]:
        if cell.value == int(user_id):
            user = worksheet[cell.row]
            print([cell.value for cell in user])
            return user, cell.row, workbook

# 3. 更新用户用户 完赛后更新用户数据
def after_match_update_user_info(user_id, database, total_games, total_win_gmames, total_income):
    # 获取到用户
    user, row, workbook = get_user_info(user_id,database)
    worksheet = workbook.worksheets[1]
    worksheet.cell(row,5).value += 1
    worksheet.cell(row, 6).value += total_games
    worksheet.cell(row, 7).value += total_win_gmames
    worksheet.cell(row, 8).value += total_income
    workbook.save(database)
    workbook.close()


#123123
# database = initialize_database()
# register(database)
# get_user_info(1,database)
# after_match_update_user_info(1,database,100,50,10000)
# after_match_update_user_info(1,database,200,30,-1000)
